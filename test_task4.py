from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
import pytest
import openpyxl

#Function to get Test Data from Excel
def getExcelData(excelFile):
    spreadSheet = openpyxl.load_workbook(excelFile)
    sheetData = spreadSheet.active
    maxCol = sheetData.max_column
    maxRow = sheetData.max_row
    excelData = []
    for i in range(2, maxRow + 1):
        row = ()
        for j in range(1, maxCol + 1):
            cellData = sheetData.cell(row = i, column = j)
            row += (cellData.value,)
        excelData.append(row)
    spreadSheet.close()
    return excelData

#Function to update result in the excel sheet
def writeResultInExcel(excelFile, name, errorCount, actualResult):
    spreadSheet = openpyxl.load_workbook(excelFile)
    sheetData = spreadSheet.active
    maxRow = sheetData.max_row
    for i in range(2, maxRow + 1):
        val = str(i)
        nameData = sheetData['A' + val]
        actualCount = sheetData['M' + val]
        result = sheetData['N' + val]
        if xstr(nameData.value) == name:
             actualCount.value = errorCount
             result.value = actualResult
        spreadSheet.save(excelFile)

#Function to replace None to empty string
def xstr(s):
    return '' if s is None else str(s)

#Function to replace None to empty string fpr int
def xint(s):
    return '' if s is None else int(s)

#Sending multiple test cases from excel as Parameters
@pytest.mark.parametrize("name, email, gender, mobile, DOB, subject, hobbies, picture, address, state, city, expected, actual, msg", getExcelData("D:\Selenium Practices\Week 4\Test Form 2.xlsx"))
#Function for sending test data to the form
def test_submit(setup_browser, name, email, gender, mobile, DOB, subject, hobbies, picture, address, state, city, expected, actual, msg):
    driver = setup_browser #this one can be retrieved from pytest fixture
    driver.get("https://www.tutorialspoint.com/selenium/practice/selenium_automation_practice.php")
    driver.implicitly_wait(30)
    actions = ActionChains(driver)
        
    #Sending the test data to the form
    driver.find_element(By.NAME, "name").send_keys(xstr(name))
    driver.find_element(By.NAME, "email").send_keys(xstr(email))  
    driver.find_element(By.NAME, "mobile").send_keys(xstr(xint(mobile)))
    driver.find_element(By.NAME, "dob").click()
    driver.find_element(By.NAME, "dob").send_keys(xstr(DOB))
    driver.find_element(By.NAME, "subjects").send_keys(xstr(subject))
    driver.find_element(By.XPATH, "//input[@name='picture']").send_keys("D:\Selenium Practices\Week 1\wikipedia_selenium_search.png")
    driver.find_element(By.XPATH, "//textarea[@name='picture']").send_keys(xstr(address))
    driver.find_element(By.XPATH, "//select[@name='state']/option[@value='"+ state +"']").click
    driver.find_element(By.XPATH, "//select[@name='city']/option[@value='"+ city +"']").click()
    driver.implicitly_wait(30)
    actions.move_to_element(driver.find_element(By.XPATH, "//input[@type='submit']")).perform()
    driver.find_element(By.XPATH, "//input[@type='submit']").click()
    try:    
        #verifying if success message is present
        errors = driver.find_elements(By.XPATH, "//label[contains(@class, 'error')]")
        errorFields = []
        for error in errors:
            errorField = error.get_attribute('for')
            errorFields.append(xstr(errorField))
        errorFields.append("are required")
        writeResultInExcel("D:\Selenium Practices\Week 4\Test Form 2.xlsx", xstr(name), xstr(len(errors)), " ".join(errorFields))
    except:
        errors = []
        actualResult = "No errors are there"
        writeResultInExcel("D:\Selenium Practices\Week 4\Test Form 2.xlsx", xstr(name), xstr(len(errors)), actualResult)
    finally:
        assert len(errors) == int(expected) #Verifying if expected no of errors matches
    
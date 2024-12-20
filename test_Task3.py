from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
import pytest
import openpyxl
import time

#Function to get Test Data from Excel
def getExcelData(excelFile):
    spreadSheet = openpyxl.load_workbook(excelFile)
    sheetData = spreadSheet.active
    maxCol = sheetData.max_column
    maxRow = sheetData.max_row
    excelData = []
    for i in range(2, maxRow + 1):
        row = ()
        for j in range(1, maxCol + 1):
            cellData = sheetData.cell(row = i, column = j)
            row += (cellData.value,)
        excelData.append(row)
    spreadSheet.close()
    return excelData

def writeResultInExcel(excelFile, name, actualResult):
    spreadSheet = openpyxl.load_workbook(excelFile)
    sheetData = spreadSheet.active
    maxRow = sheetData.max_row
    for i in range(2, maxRow + 1):
        val = str(i)
        nameData = sheetData['A' + val]
        result = sheetData['J' + val]
        if xstr(nameData.value) == name:
             result.value = actualResult
        spreadSheet.save(excelFile)

#Function to replace None to empty string
def xstr(s):
    return '' if s is None else str(s)

#Sending multiple test cases from excel as Parameters
@pytest.mark.parametrize("name, email, mobile, gender, expYears, skills, qaTools, details, expectedMessage, result", getExcelData("D:\Selenium Practices\Week 4\Test Form.xlsx"))
#Function for sending test data to the form
def test_submit(setup_browser, name, email, mobile, gender, expYears, skills, qaTools, details, expectedMessage, result):
    driver = setup_browser #this one can be retrieved from pytest fixture
    driver.get("https://qavalidation.com/demo-form/")
    driver.implicitly_wait(30)
    actions = ActionChains(driver)
    
    #Sending the test data to the form
    driver.find_element(By.NAME, "g4072-fullname").send_keys(xstr(name))
    driver.find_element(By.NAME, "g4072-email").send_keys(xstr(email))  
    driver.find_element(By.NAME, "g4072-phonenumber").send_keys(mobile)
    driver.find_element(By.XPATH, "//select[@name='g4072-gender']/option[text()='"+ gender +"']").click()
    experience = "Above 5" if int(expYears) > 5 else str(expYears)
    driver.find_element(By.ID, "g4072-yearsofexperience-" + experience).click()
    qaSkills = skills.split(', ')
    for skill in qaSkills:
        driver.find_element(By.ID, "g4072-skills-" + skill).click()
    driver.find_element(By.XPATH, "//select[@name='g4072-qatools']/option[text()='"+ qaTools +"']").click()
    driver.find_element(By.NAME, "g4072-otherdetails").send_keys(xstr(details))
    driver.implicitly_wait(30)
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    try:
        
        #verifying if success message is present
        successMessage = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,"//h4[@id='contact-form-success-header']"))
        )
        actualMessage = "Should Pass"
        actualResult = "Passed as expected"
        writeResultInExcel("D:\Selenium Practices\Week 4\Test Form.xlsx", xstr(name), actualResult)
    except:
        actualMessage = "Should Fail"
        actualResult = "Failed as expected"
        writeResultInExcel("D:\Selenium Practices\Week 4\Test Form.xlsx", xstr(name), actualResult)
    finally:
        assert expectedMessage == actualMessage #Verifying if expected scenario occurs
    
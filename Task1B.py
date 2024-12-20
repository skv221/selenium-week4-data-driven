import openpyxl
import  re #For checking valid mail
from prettytable import PrettyTable #For better visualisation

ss = openpyxl.load_workbook("D:\Selenium Practices\Week 4\Test Form.xlsx")

sheetData = ss.active

maxCol = sheetData.max_column
maxRow = sheetData.max_row

regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b' #This regex explains valid mail address

title = []

#Taking titles to print them as table
for i in range(1, maxCol + 1):
    cellObj = sheetData.cell(row=1, column=i)
    title.append(cellObj.value) 
    
table = PrettyTable(title) #Table initialisation

rowCount = 2
for i in range(2, maxRow + 1):
    #getting name and mail values to check if they exist and valid
    val = str(i)
    name = sheetData['A' + val]
    mail = sheetData['B' + val]
    #Checks if the mail is valid
    validMail = re.match(regex, str(mail.value))
    status = sheetData['I'+val]
    #Updating the sheet based on name and mail values
    if name.value is None or validMail is None:
        status.value = "Should Fail"
    else:
        status.value = "Should Pass"
    ss.save("D:\Selenium Practices\Week 4\Test Form.xlsx")
    
    #Saving the data in table
    value = []
    for j in range(1, maxCol + 1):
        cellObj = sheetData.cell(row=i, column=j)
        value.append(cellObj.value)
    table.add_row(value)
ss.close()

print(table) #Printing the table